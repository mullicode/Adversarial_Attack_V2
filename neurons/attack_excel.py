"""Excel export helpers for attack metrics."""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

SHEET_ORDER = (
    "task_meta",
    "validator_snapshots",
    "topk_rounds",
    "topk_competitors",
    "beam_paths",
    "attack_batches",
    "regions",
    "accepted_groups",
    "flip_events",
    "prune_events",
    "roundtrip",
    "all_events",
)

PREFERRED_COLUMNS: dict[str, list[str]] = {
    "task_meta": [
        "task_id",
        "true_idx",
        "true_label",
        "min_delta",
        "epsilon",
        "effective_max_delta",
        "attack_k",
        "beam_width",
        "top_regions",
        "region_grow_initial_batch",
        "region_grow_max_batch",
        "attack_preset",
        "exported_at_utc",
    ],
    "validator_snapshots": [
        "phase",
        "task_id",
        "true_idx",
        "true_label",
        "pred_idx",
        "pred_label",
        "best_other_idx",
        "best_other_label",
        "true_logit",
        "best_other_logit",
        "untargeted_gap",
        "margin",
        "flipped",
        "norm_linf",
        "rmse",
        "changed_pixel_channels",
        "changed_rgb_pixels",
        "ssim",
        "psnr_db",
        "response_time_ms",
        "pgd_fallback",
        "accepted_groups",
        "elapsed_ms",
        "remaining_ms",
    ],
    "topk_rounds": [
        "task_id",
        "round_id",
        "true_idx",
        "true_logit",
        "competitor_switched",
        "old_competitor_idx",
        "new_competitor_idx",
        "topk_count",
        "elapsed_ms",
    ],
    "topk_competitors": [
        "task_id",
        "round_id",
        "rank",
        "competitor_idx",
        "competitor_label",
        "gap_k",
        "logit",
        "elapsed_ms",
    ],
    "beam_paths": [
        "task_id",
        "round_id",
        "beam_rank",
        "beam_id",
        "untargeted_gap",
        "changed_pixels",
        "rmse",
        "gain_per_pixel",
        "flipped",
        "norm_linf",
        "elapsed_ms",
    ],
    "attack_batches": [
        "task_id",
        "round_id",
        "beam_id",
        "competitor_idx",
        "competitor_label",
        "gap_before",
        "gap_after",
        "real_gain",
        "gain_per_pixel",
        "region_yf",
        "region_xf",
        "box_y1",
        "box_y2",
        "box_x1",
        "box_x2",
        "cam_score",
        "activation_grad_score",
        "pixel_grad_density",
        "region_score",
        "batch_size",
        "accepted",
        "reason",
        "pixels",
        "norm_linf",
        "rmse",
        "pred_idx",
        "best_other_idx",
        "margin",
        "flipped",
        "elapsed_ms",
        "remaining_ms",
    ],
    "regions": [
        "task_id",
        "round_id",
        "competitor_idx",
        "region_yf",
        "region_xf",
        "pixels_applied",
        "region_fail_count",
        "region_weak_count",
        "saturated",
        "stopped_reason",
        "elapsed_ms",
    ],
    "accepted_groups": [
        "task_id",
        "group_id",
        "competitor_idx",
        "region_yf",
        "region_xf",
        "pixels",
        "gain",
        "gain_per_pixel",
        "gap_before",
        "gap_after",
        "round_id",
    ],
    "flip_events": [
        "task_id",
        "pred_idx",
        "pred_label",
        "true_idx",
        "true_label",
        "margin",
        "norm_linf",
        "rmse",
        "ssim",
        "psnr_db",
        "changed_pixel_channels",
        "elapsed_ms",
    ],
    "prune_events": [
        "task_id",
        "before_pixels",
        "after_pixels",
        "removed_pixels",
        "before_rmse",
        "after_rmse",
        "margin",
        "pred_idx",
        "flipped",
        "roundtrip_ok",
        "accepted_groups",
        "elapsed_ms",
    ],
    "roundtrip": [
        "task_id",
        "roundtrip_pred",
        "roundtrip_label",
        "roundtrip_flipped",
        "roundtrip_norm",
        "roundtrip_rmse",
        "roundtrip_margin",
        "roundtrip_ok",
        "restored_from_backup",
        "reason",
        "ssim",
        "psnr_db",
        "min_delta",
        "effective_max_delta",
        "elapsed_ms",
    ],
    "all_events": [
        "event_type",
        "task_id",
        "recorded_at_ms",
        "round_id",
        "beam_id",
        "competitor_idx",
        "gap_before",
        "gap_after",
        "real_gain",
        "gain_per_pixel",
        "accepted",
        "flipped",
        "norm_linf",
        "rmse",
        "margin",
        "pred_idx",
        "elapsed_ms",
    ],
}


def sanitize_filename(value: str, *, fallback: str = "task") -> str:
    cleaned = re.sub(r"[^\w.\-]+", "_", str(value).strip())
    cleaned = cleaned.strip("._")
    return cleaned[:120] if cleaned else fallback


def _column_order(sheet_name: str, rows: list[dict[str, Any]]) -> list[str]:
    preferred = list(PREFERRED_COLUMNS.get(sheet_name, []))
    for row in rows:
        for key in row:
            if key not in preferred:
                preferred.append(key)
    return preferred


def _write_sheet(workbook, sheet_name: str, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    worksheet = workbook.create_sheet(title=sheet_name[:31])
    columns = _column_order(sheet_name, rows)
    worksheet.append(columns)
    for row in rows:
        worksheet.append([row.get(column) for column in columns])


def write_attack_workbook(
    *,
    output_path: Path,
    sheets: dict[str, list[dict[str, Any]]],
) -> Path:
    from openpyxl import Workbook

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name in SHEET_ORDER:
        _write_sheet(workbook, sheet_name, sheets.get(sheet_name, []))

    for sheet_name, rows in sheets.items():
        if sheet_name not in SHEET_ORDER and rows:
            _write_sheet(workbook, sheet_name, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def append_summary_row(summary_path: Path, row: dict[str, Any]) -> Path:
    from openpyxl import Workbook, load_workbook

    summary_path.parent.mkdir(parents=True, exist_ok=True)
    columns = list(PREFERRED_COLUMNS["validator_snapshots"])

    if summary_path.exists():
        workbook = load_workbook(summary_path)
        worksheet = workbook.active
        existing_header = [cell.value for cell in worksheet[1]] if worksheet.max_row >= 1 else []
        columns = list(existing_header) if existing_header else columns
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "all_tasks_summary"

    for key in row:
        if key not in columns:
            columns.append(key)

    if worksheet.max_row == 0 or [cell.value for cell in worksheet[1]] != columns:
        existing_rows: list[list[Any]] = []
        if worksheet.max_row >= 2:
            old_header = [cell.value for cell in worksheet[1]]
            for row_idx in range(2, worksheet.max_row + 1):
                existing_rows.append([worksheet.cell(row=row_idx, column=col_idx + 1).value for col_idx in range(len(old_header))])
        worksheet.delete_rows(1, worksheet.max_row)
        worksheet.append(columns)
        for existing_row in existing_rows:
            remapped = {old_header[idx]: existing_row[idx] for idx in range(len(old_header))}
            worksheet.append([remapped.get(column) for column in columns])

    worksheet.append([row.get(column) for column in columns])
    workbook.save(summary_path)
    return summary_path
